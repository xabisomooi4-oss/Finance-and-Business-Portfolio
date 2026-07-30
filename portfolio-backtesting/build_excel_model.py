"""Builds a real, formula-driven Excel model for the Portfolio Analysis &
Backtesting Model, seeded with a live yfinance monthly-price snapshot.
Every calculated cell is an Excel formula, not a pasted-in Python result --
change a weight or the risk-free rate and the whole model recalculates.

Monthly (not daily) granularity is used for the return series -- standard
practice for a cross-asset correlation/risk exhibit, and it keeps the sheet
a reviewable size. The portfolio return each month is computed against the
*target* weights (not drifted ones), which is a close, standard approximation
of quarterly-rebalance-to-target for a summary exhibit like this one.

Run standalone: python build_excel_model.py
"""

from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from data import HOLDINGS, BENCHMARK

# ---------------------------------------------------------------- styling --
FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")             # hardcoded inputs
BLACK = Font(name=FONT, size=10, color="000000")            # formulas
GREEN = Font(name=FONT, size=10, color="008000")            # cross-sheet links
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
SECTION = Font(name=FONT, size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="1F2937")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="B7B7B7")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD2 = '$#,##0.00;($#,##0.00);"-"'
USD0 = '$#,##0;($#,##0);"-"'
PCT1 = '0.0%;(0.0%);"-"'
PCT2 = '0.00%;(0.00%);"-"'
NUM2 = '0.00;(0.00);"-"'


def section(ws, row, col, text, span=6):
    ws.cell(row=row, column=col, value=text).font = SECTION
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = SECTION_FILL


def label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD if bold else Font(name=FONT, size=10)
    return cell


def note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT, size=8, italic=True, color="808080")
    return cell


TICKERS = HOLDINGS + [BENCHMARK]  # NVDA..TLT, then SPY -- fixed column order throughout
COLS = {t: get_column_letter(2 + i) for i, t in enumerate(TICKERS)}  # B..J
TECH = [t for t in HOLDINGS if t not in ("GLD", "TLT")]

# ------------------------------------------------------------- fetch data --
print("Loading monthly price snapshot...")
prices = pd.read_csv("tableau_data/monthly_prices.csv", parse_dates=["Date"])
prices = prices[["Date"] + TICKERS]
n_months = len(prices)
today = date.today().isoformat()

RISK_FREE_RATE = 0.046
STARTING_VALUE = 10_000

wb = Workbook()

# ============================================================ ASSUMPTIONS =
ws = wb.active
ws.title = "Assumptions"
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 55

ws["A1"] = "Portfolio Analysis & Backtesting Model"
ws["A1"].font = TITLE
ws["A2"] = f"Snapshot date: {today}"
ws["A2"].font = Font(name=FONT, size=9, italic=True)
ws["A3"] = "Legend: blue = hardcoded input, black = formula, green = link to another sheet, yellow = key editable assumption"
ws["A3"].font = Font(name=FONT, size=8, italic=True, color="808080")

r = 5
section(ws, r, 1, "HOLDINGS & TARGET WEIGHTS", span=3)
r += 1
c = ws.cell(row=r, column=1, value="Ticker")
c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
c = ws.cell(row=r, column=2, value="Target Weight")
c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
r += 1

weight_row = {}
first_weight_row = r
for t in HOLDINGS:
    label(ws, r, 1, t)
    c = ws.cell(row=r, column=2, value=1 / len(HOLDINGS))
    c.font = BLUE
    c.number_format = PCT1
    c.border = BOX
    weight_row[t] = r
    r += 1
last_weight_row = r - 1

label(ws, r, 1, "Total", bold=True)
c = ws.cell(row=r, column=2, value=f"=SUM(B{first_weight_row}:B{last_weight_row})")
c.font = BOLD
c.number_format = PCT1
c.border = BOX
note(ws, r, 3, "Sanity check -- must equal 100.0%")
r += 2

section(ws, r, 1, "BENCHMARK, PERIOD & RATES", span=3)
r += 1
label(ws, r, 1, "Benchmark")
c = ws.cell(row=r, column=2, value=BENCHMARK)
c.font = BLUE
c.border = BOX
r += 1
label(ws, r, 1, "Period")
c = ws.cell(row=r, column=2, value=f"3 Years ({n_months} monthly closes)")
c.font = BLUE
c.border = BOX
r += 1
label(ws, r, 1, "Risk-Free Rate (10y US Treasury)")
c = ws.cell(row=r, column=2, value=RISK_FREE_RATE)
c.font = BLUE
c.number_format = PCT1
c.fill = YELLOW_FILL
c.border = BOX
note(ws, r, 3, "Live: yfinance ^TNX -- edit to test Sharpe sensitivity")
rf_row = r
r += 1
label(ws, r, 1, "Starting Portfolio Value")
c = ws.cell(row=r, column=2, value=STARTING_VALUE)
c.font = BLUE
c.number_format = USD0
c.fill = YELLOW_FILL
c.border = BOX
starting_value_row = r
r += 1

print("Assumptions sheet built.")

# ========================================================= MONTHLY RETURNS =
ws2 = wb.create_sheet("Monthly Returns")
ws2.column_dimensions["A"].width = 12
for t in TICKERS:
    ws2.column_dimensions[COLS[t]].width = 11
for col in "KLMNOPQ":
    ws2.column_dimensions[col].width = 15

ws2["A1"] = "Monthly Prices & Returns"
ws2["A1"].font = TITLE
ws2["A2"] = "Prices: yfinance monthly close, auto-adjusted for splits/dividends. Portfolio return each month is computed against target weights (rebalance-to-target approximation of the quarterly rebalance)."
ws2["A2"].font = Font(name=FONT, size=8, italic=True, color="808080")
ws2.merge_cells("A2:J2")

r = 4
c = ws2.cell(row=r, column=1, value="Date")
c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
for t in TICKERS:
    c = ws2.cell(row=r, column=2 + TICKERS.index(t), value=t)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
price_header_row = r
r += 1
price_first_row = r
for _, row_data in prices.iterrows():
    c = ws2.cell(row=r, column=1, value=row_data["Date"].strftime("%Y-%m-%d"))
    c.font = BLUE
    c.border = BOX
    for t in TICKERS:
        c = ws2.cell(row=r, column=2 + TICKERS.index(t), value=round(float(row_data[t]), 4))
        c.font = BLUE
        c.number_format = USD2
        c.border = BOX
    r += 1
price_last_row = r - 1
r += 1

# weight row, linked from Assumptions, aligned to the same B..I columns as the holdings
weights_link_row = r
label(ws2, r, 1, "Weights →", bold=True)
for t in HOLDINGS:
    c = ws2.cell(row=r, column=2 + TICKERS.index(t), value=f"=Assumptions!B{weight_row[t]}")
    c.font = GREEN
    c.number_format = PCT1
r += 2

headers = ["Date"] + TICKERS + ["Portfolio", "Port. Equity", "Port. Running Max", "Port. Drawdown",
                                 "SPY Equity", "SPY Running Max", "SPY Drawdown"]
for i, h in enumerate(headers):
    c = ws2.cell(row=r, column=1 + i, value=h)
    c.font = BOLD
    c.fill = HEADER_FILL
    c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="center")
ret_header_row = r
r += 1
ret_first_row = r

PORT_COL = "K"
PORT_EQ_COL = "L"
PORT_MAX_COL = "M"
PORT_DD_COL = "N"
SPY_EQ_COL = "O"
SPY_MAX_COL = "P"
SPY_DD_COL = "Q"

for i in range(n_months - 1):
    cur_price_row = price_first_row + 1 + i
    prev_price_row = price_first_row + i

    ws2.cell(row=r, column=1, value=f"=A{cur_price_row}").font = BLACK
    for t in TICKERS:
        col = COLS[t]
        c = ws2.cell(row=r, column=2 + TICKERS.index(t), value=f"={col}{cur_price_row}/{col}{prev_price_row}-1")
        c.font = BLACK
        c.number_format = PCT2
        c.border = BOX

    c = ws2.cell(row=r, column=11, value=f"=SUMPRODUCT($B${weights_link_row}:$I${weights_link_row},B{r}:I{r})")
    c.font = BLACK
    c.number_format = PCT2
    c.border = BOX

    spy_col = COLS[BENCHMARK]
    if i == 0:
        eq_formula = f"=Assumptions!$B${starting_value_row}*(1+{PORT_COL}{r})"
        spy_eq_formula = f"=Assumptions!$B${starting_value_row}*(1+{spy_col}{r})"
    else:
        eq_formula = f"={PORT_EQ_COL}{r - 1}*(1+{PORT_COL}{r})"
        spy_eq_formula = f"={SPY_EQ_COL}{r - 1}*(1+{spy_col}{r})"
    c = ws2.cell(row=r, column=12, value=eq_formula)
    c.font = BLACK; c.number_format = USD0; c.border = BOX

    c = ws2.cell(row=r, column=13, value=f"=MAX(${PORT_EQ_COL}${ret_first_row}:{PORT_EQ_COL}{r})")
    c.font = BLACK; c.number_format = USD0; c.border = BOX
    c = ws2.cell(row=r, column=14, value=f"={PORT_EQ_COL}{r}/{PORT_MAX_COL}{r}-1")
    c.font = BLACK; c.number_format = PCT1; c.border = BOX

    c = ws2.cell(row=r, column=15, value=spy_eq_formula)
    c.font = BLACK; c.number_format = USD0; c.border = BOX
    c = ws2.cell(row=r, column=16, value=f"=MAX(${SPY_EQ_COL}${ret_first_row}:{SPY_EQ_COL}{r})")
    c.font = BLACK; c.number_format = USD0; c.border = BOX
    c = ws2.cell(row=r, column=17, value=f"={SPY_EQ_COL}{r}/{SPY_MAX_COL}{r}-1")
    c.font = BLACK; c.number_format = PCT1; c.border = BOX

    r += 1
ret_last_row = r - 1

print("Monthly Returns sheet built.")

# ======================================================= CORRELATION MATRIX =
ws3 = wb.create_sheet("Correlation Matrix")
ws3.column_dimensions["A"].width = 10
for t in HOLDINGS:
    ws3.column_dimensions[COLS[t]].width = 10

ws3["A1"] = "Correlation Matrix (8 Holdings)"
ws3["A1"].font = TITLE
ws3["A2"] = "Pairwise Pearson correlation of monthly returns -- the actual diagnostic for whether 8 positions means real diversification."
ws3["A2"].font = Font(name=FONT, size=9, italic=True)

r = 4
for t in HOLDINGS:
    c = ws3.cell(row=r, column=2 + HOLDINGS.index(t), value=t)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
corr_header_row = r
r += 1
corr_first_row = r
mret_sheet = "'Monthly Returns'"
for t_row in HOLDINGS:
    c = ws3.cell(row=r, column=1, value=t_row)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
    col_i = COLS[t_row]
    for t_col in HOLDINGS:
        col_j = COLS[t_col]
        formula = f"=CORREL({mret_sheet}!{col_i}${ret_first_row}:{col_i}${ret_last_row},{mret_sheet}!{col_j}${ret_first_row}:{col_j}${ret_last_row})"
        c = ws3.cell(row=r, column=2 + HOLDINGS.index(t_col), value=formula)
        c.font = BLACK
        c.number_format = NUM2
        c.border = BOX
    r += 1
corr_last_row = r - 1

matrix_range = f"B{corr_first_row}:I{corr_last_row}"
ws3.conditional_formatting.add(
    matrix_range,
    ColorScaleRule(
        start_type="num", start_value=-0.2, start_color="4B8EC4",
        mid_type="num", mid_value=0.4, mid_color="F2F2F2",
        end_type="num", end_value=1, end_color="BC8434",
    ),
)

r += 1
n = len(HOLDINGS)
label(ws3, r, 1, "Average Pairwise Correlation (all 8)", bold=True)
c = ws3.cell(row=r, column=2, value=f"=(SUM({matrix_range})-{n})/{n * (n - 1)}")
c.font = BLACK
c.number_format = NUM2
c.fill = YELLOW_FILL
c.border = BOX
r += 1

tech_range = f"B{corr_first_row}:{COLS[TECH[-1]]}{corr_first_row + len(TECH) - 1}"
label(ws3, r, 1, "Average Pairwise Correlation (tech only)", bold=True)
c = ws3.cell(row=r, column=2, value=f"=(SUM({tech_range})-{len(TECH)})/{len(TECH) * (len(TECH) - 1)}")
c.font = BLACK
c.number_format = NUM2
c.fill = YELLOW_FILL
c.border = BOX
note(ws3, r, 3, f"{'/'.join(TECH)} only -- excludes GLD/TLT")
r += 1

print("Correlation Matrix sheet built.")

# ============================================================ RISK METRICS =
ws4 = wb.create_sheet("Risk Metrics")
ws4.column_dimensions["A"].width = 22
for col in "BCDEFGHIJK":
    ws4.column_dimensions[col].width = 12

ws4["A1"] = "Portfolio Risk Metrics"
ws4["A1"].font = TITLE
ws4["A2"] = "Portfolio volatility computed via the full covariance matrix (w' Σ w), not a weighted average -- the gap vs. the naive number below IS the diversification benefit, quantified."
ws4["A2"].font = Font(name=FONT, size=9, italic=True)
ws4.merge_cells("A2:F2")
ws4.row_dimensions[2].height = 26

r = 4
section(ws4, r, 1, "INDIVIDUAL ASSET STATISTICS (annualized)", span=3)
r += 1
for h, fmt in [("Ticker", None), ("Annualized Return", None), ("Annualized Volatility", None)]:
    c = ws4.cell(row=r, column=1 + ["Ticker", "Annualized Return", "Annualized Volatility"].index(h), value=h)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
asset_header_row = r
r += 1
asset_stat_row = {}
for t in HOLDINGS:
    col = COLS[t]
    c = ws4.cell(row=r, column=1, value=t)
    c.font = BOLD; c.border = BOX
    c = ws4.cell(row=r, column=2, value=f"=AVERAGE({mret_sheet}!{col}${ret_first_row}:{col}${ret_last_row})*12")
    c.font = BLACK; c.number_format = PCT1; c.border = BOX
    c = ws4.cell(row=r, column=3, value=f"=STDEVP({mret_sheet}!{col}${ret_first_row}:{col}${ret_last_row})*SQRT(12)")
    c.font = BLACK; c.number_format = PCT1; c.border = BOX
    asset_stat_row[t] = r
    r += 1
r += 1

section(ws4, r, 1, "COVARIANCE MATRIX (annualized)", span=3)
r += 1
for t in HOLDINGS:
    c = ws4.cell(row=r, column=2 + HOLDINGS.index(t), value=t)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
cov_header_row = r
r += 1
cov_first_row = r
for t_row in HOLDINGS:
    col_i = COLS[t_row]
    c = ws4.cell(row=r, column=1, value=t_row)
    c.font = BOLD; c.border = BOX
    for t_col in HOLDINGS:
        col_j = COLS[t_col]
        formula = f"=COVAR({mret_sheet}!{col_i}${ret_first_row}:{col_i}${ret_last_row},{mret_sheet}!{col_j}${ret_first_row}:{col_j}${ret_last_row})*12"
        c = ws4.cell(row=r, column=2 + HOLDINGS.index(t_col), value=formula)
        c.font = BLACK
        c.number_format = '0.0000'
        c.border = BOX
    r += 1
cov_last_row = r - 1
r += 1

section(ws4, r, 1, "PORTFOLIO VARIANCE  (w' Σ w)", span=3)
r += 1
for t in HOLDINGS:
    c = ws4.cell(row=r, column=2 + HOLDINGS.index(t), value=f"=Assumptions!B{weight_row[t]}")
    c.font = GREEN
    c.number_format = PCT1
wvar_weight_row = r
r += 1
wvar_first_row = r
for i, t_row in enumerate(HOLDINGS):
    c = ws4.cell(row=r, column=1, value=t_row)
    c.font = BOLD; c.border = BOX
    c = ws4.cell(row=r, column=11, value=f"=Assumptions!B{weight_row[t_row]}")
    c.font = GREEN
    c.number_format = PCT1
    cov_row = cov_first_row + i
    for t_col in HOLDINGS:
        col_j = COLS[t_col]
        formula = f"=$K{r}*{col_j}${wvar_weight_row}*{col_j}{cov_row}"
        c = ws4.cell(row=r, column=2 + HOLDINGS.index(t_col), value=formula)
        c.font = BLACK
        c.number_format = '0.000000'
        c.border = BOX
    r += 1
wvar_last_row = r - 1
r += 1

label(ws4, r, 1, "Portfolio Variance (annualized)", bold=True)
c = ws4.cell(row=r, column=2, value=f"=SUM(B{wvar_first_row}:I{wvar_last_row})")
c.font = BLACK; c.number_format = '0.0000'; c.border = BOX
port_var_row = r
r += 1
label(ws4, r, 1, "Portfolio Volatility -- Real (covariance-based)", bold=True)
c = ws4.cell(row=r, column=2, value=f"=SQRT(B{port_var_row})")
c.font = BLACK; c.number_format = PCT1; c.fill = YELLOW_FILL; c.border = BOX
port_vol_row = r
r += 1
label(ws4, r, 1, "Naive Weighted-Average Volatility", bold=True)
c = ws4.cell(
    row=r, column=2,
    value=f"=SUMPRODUCT(Assumptions!B{first_weight_row}:B{last_weight_row},C{asset_header_row + 1}:C{asset_header_row + len(HOLDINGS)})",
)
c.font = BLACK; c.number_format = PCT1; c.fill = YELLOW_FILL; c.border = BOX
naive_vol_row = r
note(ws4, r, 3, "What volatility would be with zero diversification benefit (perfectly correlated holdings)")
r += 1
label(ws4, r, 1, "Diversification Benefit", bold=True)
c = ws4.cell(row=r, column=2, value=f"=1-B{port_vol_row}/B{naive_vol_row}")
c.font = BOLD; c.number_format = PCT1; c.fill = YELLOW_FILL; c.border = BOX
div_benefit_row = r
r += 2

section(ws4, r, 1, "PORTFOLIO PERFORMANCE", span=3)
r += 1
label(ws4, r, 1, "Portfolio Annualized Return")
c = ws4.cell(row=r, column=2, value=f"=AVERAGE({mret_sheet}!{PORT_COL}${ret_first_row}:{PORT_COL}${ret_last_row})*12")
c.font = BLACK; c.number_format = PCT1; c.border = BOX
port_return_row = r
r += 1
label(ws4, r, 1, "Portfolio Sharpe Ratio", bold=True)
c = ws4.cell(row=r, column=2, value=f"=(B{port_return_row}-Assumptions!B{rf_row})/B{port_vol_row}")
c.font = BOLD; c.number_format = NUM2; c.border = BOX
port_sharpe_row = r
r += 1
label(ws4, r, 1, "Portfolio Max Drawdown", bold=True)
c = ws4.cell(row=r, column=2, value=f"=MIN({mret_sheet}!{PORT_DD_COL}${ret_first_row}:{PORT_DD_COL}${ret_last_row})")
c.font = BOLD; c.number_format = PCT1; c.border = BOX
note(ws4, r, 3, "Measured on monthly closes -- understates true peak-to-trough vs. a daily series, which can fall and partly recover within a single month")
port_dd_row = r
r += 1
label(ws4, r, 1, "Beta vs. Benchmark (SPY)")
spy_col = COLS[BENCHMARK]
c = ws4.cell(
    row=r, column=2,
    value=f"=COVAR({mret_sheet}!{PORT_COL}${ret_first_row}:{PORT_COL}${ret_last_row},{mret_sheet}!{spy_col}${ret_first_row}:{spy_col}${ret_last_row})/VARP({mret_sheet}!{spy_col}${ret_first_row}:{spy_col}${ret_last_row})",
)
c.font = BLACK; c.number_format = NUM2; c.border = BOX
beta_row = r
r += 2

section(ws4, r, 1, "BENCHMARK (SPY) PERFORMANCE", span=3)
r += 1
label(ws4, r, 1, "SPY Annualized Return")
c = ws4.cell(row=r, column=2, value=f"=AVERAGE({mret_sheet}!{spy_col}${ret_first_row}:{spy_col}${ret_last_row})*12")
c.font = BLACK; c.number_format = PCT1; c.border = BOX
spy_return_row = r
r += 1
label(ws4, r, 1, "SPY Annualized Volatility")
c = ws4.cell(row=r, column=2, value=f"=STDEVP({mret_sheet}!{spy_col}${ret_first_row}:{spy_col}${ret_last_row})*SQRT(12)")
c.font = BLACK; c.number_format = PCT1; c.border = BOX
spy_vol_row = r
r += 1
label(ws4, r, 1, "SPY Sharpe Ratio")
c = ws4.cell(row=r, column=2, value=f"=(B{spy_return_row}-Assumptions!B{rf_row})/B{spy_vol_row}")
c.font = BLACK; c.number_format = NUM2; c.border = BOX
spy_sharpe_row = r
r += 1
label(ws4, r, 1, "SPY Max Drawdown")
c = ws4.cell(row=r, column=2, value=f"=MIN({mret_sheet}!{SPY_DD_COL}${ret_first_row}:{SPY_DD_COL}${ret_last_row})")
c.font = BLACK; c.number_format = PCT1; c.border = BOX
spy_dd_row = r
r += 1

print("Risk Metrics sheet built.")

# ================================================================= SUMMARY =
ws5 = wb.create_sheet("Summary")
ws5.column_dimensions["A"].width = 30
for col in "BCDE":
    ws5.column_dimensions[col].width = 16

ws5["A1"] = "Portfolio Backtest Summary"
ws5["A1"].font = TITLE
ws5["A2"] = f"Equal-weighted, quarterly-rebalanced 8-asset portfolio vs. {BENCHMARK} -- snapshot {today}"
ws5["A2"].font = Font(name=FONT, size=9, italic=True)

r = 4
headers = ["Metric", "Portfolio", BENCHMARK]
for i, h in enumerate(headers):
    c = ws5.cell(row=r, column=1 + i, value=h)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
r += 1

summary_rows = [
    ("Annualized Return", f"='Risk Metrics'!B{port_return_row}", f"='Risk Metrics'!B{spy_return_row}", PCT1),
    ("Annualized Volatility", f"='Risk Metrics'!B{port_vol_row}", f"='Risk Metrics'!B{spy_vol_row}", PCT1),
    ("Sharpe Ratio", f"='Risk Metrics'!B{port_sharpe_row}", f"='Risk Metrics'!B{spy_sharpe_row}", NUM2),
    ("Max Drawdown", f"='Risk Metrics'!B{port_dd_row}", f"='Risk Metrics'!B{spy_dd_row}", PCT1),
]
for name, pf, bf, fmt in summary_rows:
    label(ws5, r, 1, name)
    c = ws5.cell(row=r, column=2, value=pf)
    c.font = GREEN; c.number_format = fmt; c.border = BOX
    c = ws5.cell(row=r, column=3, value=bf)
    c.font = GREEN; c.number_format = fmt; c.border = BOX
    r += 1
r += 1

label(ws5, r, 1, "Diversification Benefit", bold=True)
c = ws5.cell(row=r, column=2, value=f"='Risk Metrics'!B{div_benefit_row}")
c.font = GREEN; c.number_format = PCT1; c.fill = YELLOW_FILL; c.border = BOX
note(ws5, r, 3, "Real (covariance-based) vol vs. naive weighted-average vol")
r += 1
label(ws5, r, 1, "Beta vs. Benchmark", bold=True)
c = ws5.cell(row=r, column=2, value=f"='Risk Metrics'!B{beta_row}")
c.font = GREEN; c.number_format = NUM2; c.border = BOX
r += 1
label(ws5, r, 1, "Avg. Pairwise Correlation (all 8)", bold=True)
c = ws5.cell(row=r, column=2, value="='Correlation Matrix'!B" + str(corr_last_row + 2))
c.font = GREEN; c.number_format = NUM2; c.border = BOX
r += 1
label(ws5, r, 1, "Avg. Pairwise Correlation (tech only)", bold=True)
c = ws5.cell(row=r, column=2, value="='Correlation Matrix'!B" + str(corr_last_row + 3))
c.font = GREEN; c.number_format = NUM2; c.border = BOX
r += 3

# --- Equity curve chart ---
chart = LineChart()
chart.title = f"Portfolio vs. {BENCHMARK} -- Growth of ${STARTING_VALUE:,}"
chart.y_axis.title = "Value ($)"
chart.x_axis.title = "Date"
chart.style = 2
chart.height = 10
chart.width = 22

port_eq_ref = Reference(ws2, min_col=12, min_row=ret_header_row, max_row=ret_last_row)
spy_eq_ref = Reference(ws2, min_col=15, min_row=ret_header_row, max_row=ret_last_row)
chart.add_data(port_eq_ref, titles_from_data=True)
chart.add_data(spy_eq_ref, titles_from_data=True)
cats_ref = Reference(ws2, min_col=1, min_row=ret_first_row, max_row=ret_last_row)
chart.set_categories(cats_ref)
ws5.add_chart(chart, f"A{r}")

print("Summary sheet built.")

# =============================================================== DASHBOARD =
ws6 = wb.create_sheet("Dashboard")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions["A"].width = 2
for col in "BCDEFGHIJKLMN":
    ws6.column_dimensions[col].width = 11

ws6["B2"] = "Portfolio Dashboard"
ws6["B2"].font = TITLE
ws6["B3"] = f"Equal-weighted 8-asset portfolio vs. {BENCHMARK} — snapshot {today}"
ws6["B3"].font = Font(name=FONT, size=9, italic=True)

# --- KPI cards: label bar (gray) + big linked value (green, per the workbook's own legend) ---
kpi_defs = [
    ("Portfolio Return (Ann.)", f"='Risk Metrics'!B{port_return_row}", PCT1, "B"),
    ("Sharpe Ratio", f"='Risk Metrics'!B{port_sharpe_row}", NUM2, "E"),
    ("Diversification Benefit", f"='Risk Metrics'!B{div_benefit_row}", PCT1, "H"),
    ("Max Drawdown", f"='Risk Metrics'!B{port_dd_row}", PCT1, "K"),
]
kpi_row = 5
for kpi_label, formula, fmt, col in kpi_defs:
    ci = column_index_from_string(col)
    ws6.merge_cells(start_row=kpi_row, start_column=ci, end_row=kpi_row, end_column=ci + 1)
    lbl = ws6.cell(row=kpi_row, column=ci, value=kpi_label)
    lbl.font = Font(name=FONT, size=9, bold=True)
    lbl.fill = HEADER_FILL
    lbl.alignment = Alignment(vertical="center")
    for cc in (ci, ci + 1):
        ws6.cell(row=kpi_row, column=cc).fill = HEADER_FILL
        ws6.cell(row=kpi_row, column=cc).border = BOX

    ws6.merge_cells(start_row=kpi_row + 1, start_column=ci, end_row=kpi_row + 2, end_column=ci + 1)
    val = ws6.cell(row=kpi_row + 1, column=ci, value=formula)
    val.font = Font(name=FONT, size=20, bold=True, color="008000")
    val.number_format = fmt
    val.alignment = Alignment(vertical="center", horizontal="center")
    for rr in (kpi_row + 1, kpi_row + 2):
        for cc in (ci, ci + 1):
            ws6.cell(row=rr, column=cc).border = BOX
r = kpi_row + 4

# --- Ticker selector -> INDEX/MATCH detail panel (the interactive part) ---
section(ws6, r, 2, "HOLDING DETAIL — SELECT A TICKER", span=4)
r += 1
label(ws6, r, 2, "Ticker:", bold=True)
ticker_selector_cell = f"C{r}"
sel = ws6.cell(row=r, column=3, value=HOLDINGS[0])
sel.font = BLUE
sel.fill = YELLOW_FILL
sel.border = BOX
dv = DataValidation(type="list", formula1=f'"{",".join(HOLDINGS)}"', allow_blank=False)
dv.error = "Pick a ticker from the dropdown."
dv.prompt = "Choose a holding to see its weight, return, and volatility."
ws6.add_data_validation(dv)
dv.add(sel)
selector_row = r
r += 2

asset_first_data_row = asset_header_row + 1
asset_last_data_row = asset_header_row + len(HOLDINGS)
detail_defs = [
    ("Target Weight", f"=INDEX(Assumptions!B{first_weight_row}:B{last_weight_row},MATCH({ticker_selector_cell},Assumptions!A{first_weight_row}:A{last_weight_row},0))", PCT1),
    ("Annualized Return", f"=INDEX('Risk Metrics'!B{asset_first_data_row}:B{asset_last_data_row},MATCH({ticker_selector_cell},'Risk Metrics'!A{asset_first_data_row}:A{asset_last_data_row},0))", PCT1),
    ("Annualized Volatility", f"=INDEX('Risk Metrics'!C{asset_first_data_row}:C{asset_last_data_row},MATCH({ticker_selector_cell},'Risk Metrics'!A{asset_first_data_row}:A{asset_last_data_row},0))", PCT1),
]
for dname, dformula, dfmt in detail_defs:
    label(ws6, r, 2, dname)
    c = ws6.cell(row=r, column=3, value=dformula)
    c.font = BLACK
    c.number_format = dfmt
    c.border = BOX
    r += 1
r += 2

# --- Correlation heatmap, linked live from the Correlation Matrix sheet ---
section(ws6, r, 2, "CORRELATION MATRIX", span=9)
r += 1
heat_header_row = r
for t in HOLDINGS:
    c = ws6.cell(row=r, column=2 + HOLDINGS.index(t), value=t)
    c.font = BOLD; c.fill = HEADER_FILL; c.border = BOX
r += 1
heat_first_row = r
for t_row in HOLDINGS:
    corr_src_row = corr_first_row + HOLDINGS.index(t_row)
    lbl = ws6.cell(row=r, column=1, value=t_row)
    lbl.font = BOLD
    for t_col in HOLDINGS:
        corr_src_col = COLS[t_col]
        c = ws6.cell(row=r, column=2 + HOLDINGS.index(t_col), value=f"='Correlation Matrix'!{corr_src_col}{corr_src_row}")
        c.font = GREEN
        c.number_format = NUM2
        c.border = BOX
    r += 1
heat_last_row = r - 1
heat_range = f"B{heat_first_row}:I{heat_last_row}"
ws6.conditional_formatting.add(
    heat_range,
    ColorScaleRule(
        start_type="num", start_value=-0.2, start_color="4B8EC4",
        mid_type="num", mid_value=0.4, mid_color="F2F2F2",
        end_type="num", end_value=1, end_color="BC8434",
    ),
)
r += 2

chart_anchor_row = r

print("Dashboard KPI/selector/heatmap built.")

# --- Equity curve chart ---
eq_chart = LineChart()
eq_chart.title = f"Portfolio vs. {BENCHMARK} — Growth of ${STARTING_VALUE:,}"
eq_chart.y_axis.title = "Value ($)"
eq_chart.style = 2
eq_chart.height = 8.5
eq_chart.width = 17
port_eq_ref2 = Reference(ws2, min_col=12, min_row=ret_header_row, max_row=ret_last_row)
spy_eq_ref2 = Reference(ws2, min_col=15, min_row=ret_header_row, max_row=ret_last_row)
eq_chart.add_data(port_eq_ref2, titles_from_data=True)
eq_chart.add_data(spy_eq_ref2, titles_from_data=True)
eq_chart.set_categories(Reference(ws2, min_col=1, min_row=ret_first_row, max_row=ret_last_row))
ws6.add_chart(eq_chart, f"B{chart_anchor_row}")

# --- Return-by-holding bar chart ---
ret_chart = BarChart()
ret_chart.type = "col"
ret_chart.title = "Annualized Return by Holding"
ret_chart.style = 10
ret_chart.height = 8.5
ret_chart.width = 12
ret_data_ref = Reference(ws4, min_col=2, min_row=asset_header_row, max_row=asset_last_data_row)
ret_cats_ref = Reference(ws4, min_col=1, min_row=asset_first_data_row, max_row=asset_last_data_row)
ret_chart.add_data(ret_data_ref, titles_from_data=True)
ret_chart.set_categories(ret_cats_ref)
ret_chart.legend = None
ws6.add_chart(ret_chart, f"J{chart_anchor_row}")

# --- Allocation pie chart (all slices are equal by construction -- fine for a pie) ---
pie_chart = PieChart()
pie_chart.title = "Asset Allocation"
pie_chart.height = 8.5
pie_chart.width = 12
pie_data_ref = Reference(ws, min_col=2, min_row=first_weight_row - 1, max_row=last_weight_row)
pie_cats_ref = Reference(ws, min_col=1, min_row=first_weight_row, max_row=last_weight_row)
pie_chart.add_data(pie_data_ref, titles_from_data=True)
pie_chart.set_categories(pie_cats_ref)
ws6.add_chart(pie_chart, f"J{chart_anchor_row + 18}")

print("Dashboard charts built.")

# Move Dashboard to the first tab position
wb.move_sheet("Dashboard", offset=-(len(wb.sheetnames) - 1))
wb.active = 0

wb.save("Portfolio_Backtesting_Model.xlsx")
print("\nSaved Portfolio_Backtesting_Model.xlsx")
