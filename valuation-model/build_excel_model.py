"""Builds a real, formula-driven Excel valuation model for BLK (DDM + Comps),
seeded with a live yfinance data snapshot. Every calculated cell is an Excel
formula, not a pasted-in Python result -- change an assumption and the whole
model recalculates, same as it would on a real desk.

Run standalone: python build_excel_model.py
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from data import get_fundamentals, get_risk_free_rate
from ddm import historical_dividend_growth, EQUITY_RISK_PREMIUM, TERMINAL_GROWTH_RATE, FORECAST_YEARS
from comps import PEERS, fetch_peers

# ---------------------------------------------------------------- styling --
FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")            # hardcoded inputs
BLACK = Font(name=FONT, size=10, color="000000")            # formulas
GREEN = Font(name=FONT, size=10, color="008000")            # cross-sheet links
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True)
SECTION = Font(name=FONT, size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="1F2937")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin", color="B7B7B7")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

USD2 = '$#,##0.00;($#,##0.00);"-"'
USD0 = '$#,##0;($#,##0);"-"'
PCT1 = '0.0%;(0.0%);"-"'
MULT = '0.0"x";(0.0"x");"-"'


def section(ws, row, col, text, span=4):
    ws.cell(row=row, column=col, value=text).font = SECTION
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = SECTION_FILL


def label(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD if bold else Font(name=FONT, size=10)
    return cell


def note(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name=FONT, size=8, italic=True, color="808080")
    return cell


# ------------------------------------------------------------- fetch data --
print("Pulling live data snapshot...")
blk = get_fundamentals("BLK")
peer_symbols = PEERS
peer_data = fetch_peers(peer_symbols)
rf = get_risk_free_rate()
growth = historical_dividend_growth(blk["dividend_history"])
today = date.today().isoformat()

wb = Workbook()

# ============================================================ ASSUMPTIONS =
ws = wb.active
ws.title = "Assumptions"
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 55

ws["A1"] = "BlackRock (BLK) — Valuation Model"
ws["A1"].font = TITLE
ws["A2"] = f"Snapshot date: {today}"
ws["A2"].font = Font(name=FONT, size=9, italic=True)
ws["A3"] = "Legend: blue = hardcoded input, black = formula, green = link to another sheet, yellow = key editable assumption"
ws["A3"].font = Font(name=FONT, size=8, italic=True, color="808080")

r = 5
section(ws, r, 1, "MARKET DATA (source: yfinance, snapshot above)")
r += 1
rows_market = [
    ("Current Price", blk["price"], USD2, "yfinance Ticker.info['currentPrice']"),
    ("Shares Outstanding", blk["shares_outstanding"], '#,##0', "yfinance fast_info / balance sheet"),
    ("Total Debt", blk["total_debt"], USD0, "yfinance Ticker.info['totalDebt']"),
    ("Total Cash", blk["total_cash"], USD0, "yfinance Ticker.info['totalCash']"),
    ("Beta", blk["beta"], '0.000', "yfinance Ticker.info['beta']"),
    ("Trailing EPS", blk["trailing_eps"], USD2, "yfinance Ticker.info['trailingEps']"),
    ("EBITDA (TTM)", blk["ebitda"], USD0, "yfinance Ticker.info['ebitda']"),
    ("Current Annual Dividend", blk["trailing_annual_dividend"], USD2, "Sum of trailing 4 quarterly payments"),
]
input_rows = {}
for name, value, fmt, src in rows_market:
    label(ws, r, 1, name)
    c = ws.cell(row=r, column=2, value=value)
    c.font = BLUE
    c.number_format = fmt
    c.border = BOX
    note(ws, r, 3, f"Source: {src}")
    input_rows[name] = r
    r += 1

market_cap_row = r
label(ws, r, 1, "Market Cap", bold=True)
c = ws.cell(row=r, column=2, value=f"=B{input_rows['Current Price']}*B{input_rows['Shares Outstanding']}")
c.font = BOLD
c.number_format = USD0
c.border = BOX
r += 2

section(ws, r, 1, "RATE & GROWTH ASSUMPTIONS")
r += 1
rows_rates = [
    ("Risk-Free Rate (10y US Treasury)", rf, PCT1, "Live: yfinance ^TNX"),
    ("Equity Risk Premium", EQUITY_RISK_PREMIUM, PCT1, "Damodaran-style long-run US ERP assumption -- edit to test sensitivity"),
    ("Dividend Growth Rate (5y CAGR)", growth, PCT1, "Computed from BLK's actual quarterly dividend history, full years only"),
    ("Terminal Growth Rate", TERMINAL_GROWTH_RATE, PCT1, "Assumption: long-run sustainable growth, ~nominal GDP -- edit to test sensitivity"),
    ("Forecast Years", FORECAST_YEARS, '0', "Explicit DDM forecast horizon"),
]
for name, value, fmt, src in rows_rates:
    label(ws, r, 1, name)
    c = ws.cell(row=r, column=2, value=value)
    c.font = BLUE
    c.number_format = fmt
    c.fill = YELLOW_FILL
    c.border = BOX
    note(ws, r, 3, src)
    input_rows[name] = r
    r += 1

ASM = input_rows  # shorthand map: name -> row number, all on 'Assumptions' sheet, column B

print("Assumptions sheet built.")

# ==================================================================== DDM =
ws2 = wb.create_sheet("DDM")
ws2.column_dimensions["A"].width = 30
for col in "BCDEFGH":
    ws2.column_dimensions[col].width = 14

ws2["A1"] = "Dividend Discount Model"
ws2["A1"].font = TITLE
ws2["A2"] = "Discounts at cost of equity (CAPM), not WACC -- dividends reach equity holders specifically."
ws2["A2"].font = Font(name=FONT, size=9, italic=True)

r = 4
section(ws2, r, 1, "COST OF EQUITY (CAPM)")
r += 1
label(ws2, r, 1, "Risk-Free Rate")
c = ws2.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Risk-Free Rate (10y US Treasury)']}")
c.font = GREEN
c.number_format = PCT1
rf_row = r
r += 1
label(ws2, r, 1, "Beta")
c = ws2.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Beta']}")
c.font = GREEN
c.number_format = '0.000'
beta_row = r
r += 1
label(ws2, r, 1, "Equity Risk Premium")
c = ws2.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Equity Risk Premium']}")
c.font = GREEN
c.number_format = PCT1
erp_row = r
r += 1
label(ws2, r, 1, "Cost of Equity = Rf + Beta × ERP", bold=True)
c = ws2.cell(row=r, column=2, value=f"=B{rf_row}+B{beta_row}*B{erp_row}")
c.font = BOLD
c.number_format = PCT1
c.border = BOX
coe_row = r
r += 2

section(ws2, r, 1, "DIVIDEND PROJECTION", span=5)
r += 1
headers = ["Year", "Dividend / Share", "Discount Factor", "PV of Dividend"]
for i, h in enumerate(headers):
    c = ws2.cell(row=r, column=1 + i, value=h)
    c.font = BOLD
    c.fill = HEADER_FILL
    c.border = BOX
header_row = r
r += 1
first_div_row = r
for yr in range(1, FORECAST_YEARS + 1):
    ws2.cell(row=r, column=1, value=yr).font = BLACK
    ws2.cell(row=r, column=1).border = BOX
    if yr == 1:
        div_formula = f"=Assumptions!B{ASM['Current Annual Dividend']}*(1+Assumptions!B{ASM['Dividend Growth Rate (5y CAGR)']})"
    else:
        div_formula = f"=B{r - 1}*(1+Assumptions!B{ASM['Dividend Growth Rate (5y CAGR)']})"
    c = ws2.cell(row=r, column=2, value=div_formula)
    c.font = GREEN
    c.number_format = USD2
    c.border = BOX

    c = ws2.cell(row=r, column=3, value=f"=1/(1+$B${coe_row})^A{r}")
    c.font = BLACK
    c.number_format = '0.0000'
    c.border = BOX

    c = ws2.cell(row=r, column=4, value=f"=B{r}*C{r}")
    c.font = BLACK
    c.number_format = USD2
    c.border = BOX
    r += 1
last_div_row = r - 1
r += 1

section(ws2, r, 1, "TERMINAL VALUE (Gordon Growth Model)")
r += 1
label(ws2, r, 1, "Final Projected Dividend")
c = ws2.cell(row=r, column=2, value=f"=B{last_div_row}")
c.font = BLACK
c.number_format = USD2
r += 1
label(ws2, r, 1, "Terminal Growth Rate")
c = ws2.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Terminal Growth Rate']}")
c.font = GREEN
c.number_format = PCT1
tg_row = r
r += 1
label(ws2, r, 1, "Terminal Value = D₅×(1+g) / (Re − g)", bold=True)
c = ws2.cell(row=r, column=2, value=f"=B{r - 2}*(1+B{tg_row})/(B{coe_row}-B{tg_row})")
c.font = BOLD
c.number_format = USD0
c.border = BOX
tv_row = r
r += 1
label(ws2, r, 1, "PV of Terminal Value")
c = ws2.cell(row=r, column=2, value=f"=B{tv_row}/(1+B{coe_row})^{FORECAST_YEARS}")
c.font = BLACK
c.number_format = USD2
c.border = BOX
pv_tv_row = r
r += 2

section(ws2, r, 1, "INTRINSIC VALUE")
r += 1
label(ws2, r, 1, "Sum of PV of Dividends (Yr 1–{})".format(FORECAST_YEARS))
c = ws2.cell(row=r, column=2, value=f"=SUM(D{first_div_row}:D{last_div_row})")
c.font = BLACK
c.number_format = USD2
sum_pv_div_row = r
r += 1
label(ws2, r, 1, "PV of Terminal Value")
c = ws2.cell(row=r, column=2, value=f"=B{pv_tv_row}")
c.font = BLACK
c.number_format = USD2
r += 1
label(ws2, r, 1, "Intrinsic Value per Share", bold=True)
c = ws2.cell(row=r, column=2, value=f"=B{sum_pv_div_row}+B{pv_tv_row}")
c.font = BOLD
c.number_format = USD2
c.fill = YELLOW_FILL
c.border = BOX
intrinsic_value_row = r
r += 1
label(ws2, r, 1, "Current Market Price")
c = ws2.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Current Price']}")
c.font = GREEN
c.number_format = USD2
ddm_price_ref_row = r
r += 1
label(ws2, r, 1, "Implied Upside / (Downside)", bold=True)
c = ws2.cell(row=r, column=2, value=f"=B{intrinsic_value_row}/B{ddm_price_ref_row}-1")
c.font = BOLD
c.number_format = PCT1
r += 1

print("DDM sheet built.")

# =================================================================== COMPS =
ws3 = wb.create_sheet("Comps")
ws3.column_dimensions["A"].width = 22
for col in "BCDEFGHI":
    ws3.column_dimensions[col].width = 15

ws3["A1"] = "Comparable Company Analysis"
ws3["A1"].font = TITLE
ws3["A2"] = "Peer averages exclude 'N/A' cells automatically (AVERAGE ignores text) -- gaps are real source-data gaps, not errors."
ws3["A2"].font = Font(name=FONT, size=9, italic=True)
ws3["A3"] = (
    "Note: Enterprise Value here = Market Cap + Total Debt - Total Cash (the standard textbook formula, "
    "computed live by formula below). This is a deliberate simplification -- it omits minority interest and "
    "preferred stock, so it will differ from a data provider's own pre-calculated EV figure, which often "
    "includes those. Worth knowing: which EV convention is in use changes the implied price meaningfully."
)
ws3["A3"].font = Font(name=FONT, size=8, italic=True, color="808080")
ws3.merge_cells("A3:I3")
ws3.row_dimensions[3].height = 28

r = 4
headers = ["Ticker", "Price", "EPS", "Total Debt", "Total Cash", "EBITDA", "Enterprise Value", "P/E", "EV/EBITDA"]
for i, h in enumerate(headers):
    c = ws3.cell(row=r, column=1 + i, value=h)
    c.font = BOLD
    c.fill = HEADER_FILL
    c.border = BOX
comps_header_row = r
r += 1

all_symbols = ["BLK"] + peer_symbols
all_fund = [blk] + peer_data
comps_rows = {}
for f in all_fund:
    ws3.cell(row=r, column=1, value=f["symbol"]).font = BOLD
    ws3.cell(row=r, column=1).border = BOX

    c = ws3.cell(row=r, column=2, value=f["price"])
    c.font = BLUE
    c.number_format = USD2
    c.border = BOX

    eps_val = f["trailing_eps"] if f["trailing_eps"] is not None else "N/A"
    c = ws3.cell(row=r, column=3, value=eps_val)
    c.font = BLUE
    c.number_format = USD2
    c.border = BOX

    debt_val = f["total_debt"] if f["total_debt"] is not None else "N/A"
    c = ws3.cell(row=r, column=4, value=debt_val)
    c.font = BLUE
    c.number_format = USD0
    c.border = BOX

    cash_val = f["total_cash"] if f["total_cash"] is not None else "N/A"
    c = ws3.cell(row=r, column=5, value=cash_val)
    c.font = BLUE
    c.number_format = USD0
    c.border = BOX

    ebitda_val = f["ebitda"] if f["ebitda"] is not None else "N/A"
    c = ws3.cell(row=r, column=6, value=ebitda_val)
    c.font = BLUE
    c.number_format = USD0
    c.border = BOX

    # EV = Price*Shares (market cap) + Debt - Cash -- formula, guarded for missing debt/cash
    if f["total_debt"] is not None and f["total_cash"] is not None:
        ev_formula = f"=B{r}*{f['shares_outstanding']}+D{r}-E{r}"
        c = ws3.cell(row=r, column=7, value=ev_formula)
        c.font = BLACK
    else:
        c = ws3.cell(row=r, column=7, value="N/A")
        c.font = BLUE
    c.number_format = USD0
    c.border = BOX

    # P/E -- formula, guarded for negative/missing EPS (shows "NM" = not meaningful)
    if f["trailing_eps"] is not None and f["trailing_eps"] > 0:
        pe_formula = f'=IF(C{r}>0,B{r}/C{r},"NM")'
        c = ws3.cell(row=r, column=8, value=pe_formula)
        c.font = BLACK
    else:
        c = ws3.cell(row=r, column=8, value="NM")
        c.font = BLUE
    c.number_format = MULT
    c.border = BOX

    # EV/EBITDA -- formula, guarded for missing EBITDA/EV
    if f["ebitda"] is not None and f["total_debt"] is not None:
        c = ws3.cell(row=r, column=9, value=f"=G{r}/F{r}")
        c.font = BLACK
    else:
        c = ws3.cell(row=r, column=9, value="N/A")
        c.font = BLUE
    c.number_format = MULT
    c.border = BOX

    comps_rows[f["symbol"]] = r
    r += 1

peer_first_row = comps_rows[peer_symbols[0]]
peer_last_row = comps_rows[peer_symbols[-1]]
blk_row = comps_rows["BLK"]

r += 1
label(ws3, r, 1, "Peer Average", bold=True)
c = ws3.cell(row=r, column=8, value=f"=AVERAGE(H{peer_first_row}:H{peer_last_row})")
c.font = BOLD
c.number_format = MULT
c.fill = YELLOW_FILL
avg_pe_row = r
c = ws3.cell(row=r, column=9, value=f"=AVERAGE(I{peer_first_row}:I{peer_last_row})")
c.font = BOLD
c.number_format = MULT
c.fill = YELLOW_FILL
avg_ev_ebitda_row = r
r += 2

section(ws3, r, 1, "IMPLIED VALUE", span=3)
r += 1
label(ws3, r, 1, "Implied Price -- P/E Method")
c = ws3.cell(row=r, column=2, value=f"=H{avg_pe_row}*C{blk_row}")
c.font = BLACK
c.number_format = USD2
c.border = BOX
implied_pe_row = r
r += 1
label(ws3, r, 1, "Implied Price -- EV/EBITDA Method")
c = ws3.cell(
    row=r, column=2,
    value=f"=(I{avg_ev_ebitda_row}*F{blk_row}-D{blk_row}+E{blk_row})/{blk['shares_outstanding']}",
)
c.font = BLACK
c.number_format = USD2
c.border = BOX
implied_ev_ebitda_row = r
r += 1
label(ws3, r, 1, "Current Market Price")
c = ws3.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Current Price']}")
c.font = GREEN
c.number_format = USD2
comps_price_ref_row = r
r += 1

print("Comps sheet built.")

# ================================================================= SUMMARY =
ws4 = wb.create_sheet("Summary")
ws4.column_dimensions["A"].width = 30
for col in "BCDE":
    ws4.column_dimensions[col].width = 16

ws4["A1"] = "Valuation Summary — BLK"
ws4["A1"].font = TITLE
ws4["A2"] = f"Current Price: linked live from Assumptions — snapshot {today}"
ws4["A2"].font = Font(name=FONT, size=9, italic=True)

r = 4
headers = ["Method", "Implied Value / Share", "vs. Current Price"]
for i, h in enumerate(headers):
    c = ws4.cell(row=r, column=1 + i, value=h)
    c.font = BOLD
    c.fill = HEADER_FILL
    c.border = BOX
summary_header_row = r
r += 1

summary_rows = [
    ("DDM (dividend-based)", f"=DDM!B{intrinsic_value_row}"),
    ("Comps — P/E", f"=Comps!B{implied_pe_row}"),
    ("Comps — EV/EBITDA", f"=Comps!B{implied_ev_ebitda_row}"),
]
summary_data_first_row = r
for name, formula in summary_rows:
    label(ws4, r, 1, name)
    c = ws4.cell(row=r, column=2, value=formula)
    c.font = GREEN
    c.number_format = USD2
    c.border = BOX
    c2 = ws4.cell(row=r, column=3, value=f"=B{r}/Assumptions!B{ASM['Current Price']}-1")
    c2.font = BLACK
    c2.number_format = PCT1
    c2.border = BOX
    r += 1
summary_data_last_row = r - 1

r += 1
label(ws4, r, 1, "Current Market Price", bold=True)
c = ws4.cell(row=r, column=2, value=f"=Assumptions!B{ASM['Current Price']}")
c.font = GREEN
c.number_format = USD2
c.border = BOX
current_price_summary_row = r
r += 3

section(ws4, r, 1, "READING THE SPREAD", span=3)
r += 1
payout_note = (
    f"BLK pays out ~{blk['payout_ratio']:.0%} of earnings as dividends -- DDM only values that slice. "
    f"It cannot see value from the other ~{1 - blk['payout_ratio']:.0%}, retained and reinvested "
    f"(Aladdin, buybacks, M&A), which is most of the gap vs. its comps-implied value."
)
ws4.cell(row=r, column=1, value=payout_note).font = Font(name=FONT, size=9)
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws4.row_dimensions[r].height = 30
r += 1
comps_note = (
    f"Comps show BLK trading at a real premium to peers (own P/E {blk['trailing_pe']:.1f}x vs. "
    f"peer average) -- consistent with the market pricing in growth the DDM misses entirely, "
    f"though comps still shows skepticism about how large that premium should be."
)
ws4.cell(row=r, column=1, value=comps_note).font = Font(name=FONT, size=9)
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws4.row_dimensions[r].height = 30
r += 2

# --- Football field chart ---
chart = BarChart()
chart.type = "bar"  # horizontal
chart.title = "Valuation Football Field"
chart.y_axis.title = None
chart.x_axis.title = "Implied Value per Share ($)"
chart.style = 2

data_ref = Reference(ws4, min_col=2, min_row=summary_header_row, max_row=summary_data_last_row)
cats_ref = Reference(ws4, min_col=1, min_row=summary_data_first_row, max_row=summary_data_last_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 8
chart.width = 18
ws4.add_chart(chart, f"E{summary_header_row}")

print("Summary sheet built.")

wb.save("BLK_Valuation_Model.xlsx")
print("\nSaved BLK_Valuation_Model.xlsx")
